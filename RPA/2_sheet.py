from openpyxl import Workbook
wb = Workbook()
# wb.active 이건 활성화된 시트 가져오는거
ws = wb.create_sheet()  # 새로운 시트를 기본 이름으로 생성
ws.title = "MySheet"  # 시트 이름 변경
ws.sheet_properties.tabColor = "ff66ff"  # RGB 형태로 색상을 넣어주면 탭 컬러가 바뀜

# Sheet, MySheet, YourSheet
ws1 = wb.create_sheet("YourSheet")  # 주어진 이름으로 새로운 시트 생성
ws2 = wb.create_sheet("NewSheet", 2)  # 2번째 인덱스에 새로운 시트를 생성
# Sheet, MySheet, NewSheet, YourSheet

new_ws = wb["NewSheet"]  # 딕셔너리 형태로 시트에 접근

print(wb.sheetnames)  # 모든 시트 이름 확인하기

# Sheet 복사
new_ws["A1"] = "Test" # A1셀의 값을 Test라고 입력하는거야
target = wb.copy_worksheet(new_ws)  # 워크북 내에 있는 new_ws 워크시트를 복사해서 target이라는 워크시트를 만들어줘
target.title = "Copied Sheet"  # target시트의 이름을 "Copied Sheet"으로 바꿔줘 

wb.save("sample.xlsx")