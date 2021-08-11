from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws.title = "NadoSheet"

# 셀에 값 입력하기 1
ws["A1"] = 1  # A1 셀에 1이라는 값을 입력
ws["A2"] = 2
ws["A3"] = 3

ws["B1"] = 4
ws["B2"] = 5
ws["B3"] = 6

# 값 확인하기 1
print(ws["A1"])  # A1 셀의 정보를 출력하고 싶은데, 걍 정보만 떠 | <Cell 'NadoSheet'.A1> 
print(ws["A1"].value)   # 이렇게해야 A1 셀의 '값'을 출력 | 1
print(ws["A10"].value)  # 값이 없는 애를 출력하면 | None

# 값 확인하기 2
# row = 1, 2, 3, ...
# column = A, B, C, ... 근데, 우리는 숫자로 접근하니까 A = 1, B = 2 ... 
print(ws.cell(column=1, row=1).value)    # A1 셀을 출력 == print(ws["A1"].value) | 1
print(ws.cell(column=2, row=1).value)    # B1 셀을 출력 == print(ws["B1"].value) | 4

# 셀에 값 입력하기 2
c = ws.cell(column=3, row=1, value=10)   # ws["C1"].value = 10 이랑 똑같음
print(c.value)  # ws["C1"].value | 10



# 반복문을 이용해서 랜덤 숫자 채우기
# 먼저 random import 하기
from random import *

index = 1
for x in range(1, 11):  # 10개 row
    for y in range(1, 11):  # 10개 column
        # ws.cell(row=x, column=y, value=randint(0, 100))  # 0 ~ 100 사이의 숫자
        ws.cell(row=x, column=y, value=index)
        index += 1



wb.save("sample.xlsx")
