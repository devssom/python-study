from openpyxl import load_workbook  # 파일 불러오기
wb = load_workbook("sample.xlsx")  # sample.xlsx 파일에서 워크북을 불러옴
ws = wb.active # 활성화된 시트

# cell 데이터 불러오기
for x in range(1, 11):
    for y in range(1, 11):
        print(ws.cell(row=x, column=y).value, end=" ")  # end=" " : 셀 하나에 한 줄이 아니라, 1 2 3 4 ... 이렇게 나오게
    print()   # 줄바꿈 할 수 있도록 써준거

# cell 갯수를 모를 때
for x in range(1, ws.max_row + 1):  # 최대 row 수
    for y in range(1, ws.max_column + 1):  # 최대 column 수
        print(ws.cell(row=x, column=y).value, end=" ")  # end=" " : 셀 하나에 한 줄이 아니라, 1 2 3 4 ... 이렇게 나오게
    print()   # 줄바꿈 할 수 있도록 써준거
